from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill, Alignment
from app.main.core.io import stream_file
from app.main.config import upload_location as report_file_path
from dateutil.relativedelta import relativedelta
from datetime import datetime, timedelta
from sqlalchemy import func, and_, desc
import enum

from .apiservice import has_permissions
from app.main.model.client import Client, ClientType
from app.main.model.user import User
from app.main.model.rac import RACRole
from app.main.model.sales_board import SalesBoard
from app.main.core.rac import RACRoles

from flask import current_app as app


class ReportPeriod(enum.Enum):
    CW = 'Current Week'
    CM = 'Current Month'
    LW = 'Last Week'
    LM = 'Last Month' 
    L2M = 'Last Two Month'
    L3M = 'Last Three Month'
    L6M = 'Last Six Month'
    L12M = 'Last twelve month'

class ReportService(object):
    _permissions = []
    _heading = 'Report'
    _thead_bg_color = '0591f5'
    _report_name = 'report.xlsx'
    _report_body = []
    _start = None
    _end = None
    
    def __init__(self, period=None, start=None, end=None):
        # period
        if period:
            self._parse_period(period)
        else:
            if start:
                # date util parsing
                self._start = dt_parse(start)
            if end:
                self._end = dt_parse(end)


    def _parse_period(self, period):
        today = datetime.utcnow().date()
        if period == ReportPeriod.CW.name:
            self._start = today - timedelta(days=today.weekday()) 
        elif period == ReportPeriod.CM.name:
            self._start = today.replace(day=1)
        elif period == ReportPeriod.LW.name:
            self._start = today - timedelta(days=today.weekday()+7)
            self._end = today - timedelta(days=today.weekday()+1)
        elif period == ReportPeriod.LM.name:
            mstart = today.replace(day=1) 
            self._start = mstart - relativedelta(months=1)
            self._end   = mstart - timedelta(days=1) 
        elif period == ReportPeriod.L2M.name:
            mstart = today.replace(day=1)
            self._start = mstart - relativedelta(months=2)
            self._end   = mstart - timedelta(days=1) 
        elif period == ReportPeriod.L3M.name:
            mstart = today.replace(day=1)
            self._start = mstart - relativedelta(months=3)
            self._end   = mstart - timedelta(days=1)
        elif period == ReportPeriod.L6M.name:
            mstart = today.replace(day=1)
            self._start = mstart - relativedelta(months=6)
            self._end   = mstart - timedelta(days=1)
        elif period == ReportPeriod.L12M.name:
            mstart = today.replace(day=1)
            self._start = mstart - relativedelta(months=12)
            self._end   = mstart - timedelta(days=1)

    @has_permissions
    def view(self, data=None):
        return self._records(data)
        
    # only excel is supported
    @has_permissions
    def download(self, data=None, fmt='excel'):
        report_file = self.to_excel(data) 
        return stream_file(report_file_path, report_file, as_attachment=False)	

    @has_permissions
    def to_excel(self, data=None):
        wb = Workbook()         
        ws = wb.active        
        # fill the data
        ws['A1'] = self._heading
        ws['A3'] = 'Report Date:'
        ws['C3'] = datetime.utcnow().strftime('%m/%d/%Y')
        # styling
        ws['A1'].font = Font(bold=True)

        ws.append([])
        headers = self._fields
        headers.insert(0, '')
        ws.append(headers)
        col_head_style = PatternFill(start_color=self._thead_bg_color, end_color=self._thead_bg_color, fill_type = "solid")
        for rows in ws.iter_rows(min_row=5, max_row=5, min_col=1):
            for cell in rows:
                cell.fill = col_head_style
                cell.alignment = Alignment(wrap_text=True,vertical='top') 

        # append the rows 
        row_index = 5
        # call the abstract method the fetch the rows
        rows = self._values(data)
        for row in rows:
            row_index = row_index + 1
            ws.append(row) 
        # save the file
        report_file = "{}/{}".format(report_file_path, self._report_name)
        wb.save(filename=report_file)
        return self._report_name

    def _values(self, data):
        result = []
        records = self._records(data)
        rep = lambda s: '' if s is None else s
        index = 0
        for record in records:
            index = index + 1
            item = [rep(v) for k, v in record.items()]
            item.insert(0, index)
            result.append(item)
        return result

    @staticmethod
    def _dt2str(dt):
        if dt:
            return dt.strftime("%m/%d/%Y %I:%M %p")

        return ''

from app.main.model.address import AddressType
class ClientReportSvc(ReportService):

    @property
    def _fields(self):
        return ['Campaign Name', 'Interest Level', 'First Name', 'Last Name', 'DOB', 
                'Lead Source', 'Disposition', 'Lead Type', 'Salesman', 
                'Account Manager', 'Team Manager', 'Attorney', 'Contact #', 
                'Email', 'State', 'City', 'Zip', 'Fico', 'Smart Credit', 'Latest Action', 
                'Last Appt', 'Client ID', 'Last Call', 'Calls', 'Total Call Duration',
                'Task', 'BackEnd', 'Term', 'Days Remain', 'Commission Rate(%)',
                'Total Debt', 
               ]

    def _records(self, data):
        try:
            records = []
            filts = []
            if self._start:
                filts.append(func.date(Client.inserted_on)>=self._start)
            if self._end:
                filts.append(func.date(Client.inserted_on)<=self._end)

            filts.append(Client.type != ClientType.coclient)
            query = Client.query
            if len(filts) > 0:
                query = query.filter(and_(*filts))

            clients = query.all()
            for client in clients:
                # current address
                current_addr = None
                for addr in client.addresses:
                    if addr.type == AddressType.CURRENT:
                        current_addr = addr

                record = {
                    'campaign_name': '',
                    'interest_level': '',
                    'first_name': client.first_name,
                    'last_name': client.last_name,
                    'dob': self._dt2str(client.dob),
                    'lead_source': client.lead_source,
                    'status': client.disposition.value if client.disposition else '',
                    'lead_type': 'Primary',
                    'salesman': client.sales_rep.full_name if client.sales_rep else '',
                    'account_manager': client.account_manager.full_name if client.account_manager else '',
                    'team_manager': client.team_manager.full_name if client.team_manager else '',
                    'attorney': '', 
                    'contact': '', 
                    'email': client.email, 
                    'state': current_addr.state if current_addr else '', 
                    'city': current_addr.city if current_addr else '', 
                    'zip': current_addr.zip_code if current_addr else '', 
                    'fico': '', 
                    'smart_credit': '',
                    #'smart_credit_email': '',
                    #'smart_credit_token': '',
                    'latest_action': '',
                    'last_appt': '',
                    'client_id': client.friendly_id,
                    #'created_date': '',
                    'last_call': '',
                    #'age': '',
                    'calls': '',
                    'total_call_duration': '',
                    'task': '',
                    'backend': '',
                    'term': '',
                    'days_remain': '',
                    'commission_rate': '',
                    'total_debt': '',
                    #'next_draft_date': '',
                    #'num_payments': '',
                    #'total_payment': '',
                    #'sale_date': '',
                    #'first_payment_date': '',
                    #'balance': '',
                    #'last_update': '',
                } 
                records.append(record)

            return records
        except Exception as err:
            app.logger.warning("Client report service {}".format(str(err)))
            return []
        
from app.main.model.sales_board import SalesFlow
class SalesReportSvc(ReportService):
    _heading = 'Sales Report'
    _report_name = 'sales_report.xlsx'
   
    @property
    def _fields(self):
        return ['Agent', 'New Leads', 'New Deals', 'Recycled Leads', 'Recycled Deals', 
                'Closing % Recycled', 'Total Leads All', 'Total Closing % All', 'Retention', 
                'Total Debt' ]
        
    def _records(self, data):
        try:
            records = []
            filts = []
            if self._start:
                filts.append(func.date(Client.inserted_on)>=self._start)
            if self._end:
                filts.append(func.date(Client.inserted_on)<=self._end)
  
           
            for sb in SalesBoard.query.all():
                agent = sb.agent
                # all sales records
                total_debt = 0
                for sf in SalesFlow.query.filter_by(agent_id=agent.id).all():
                    total_debt = total_debt + sf.lead.total_debt
 
                closing_percent = round((sb.tot_deals/sb.tot_leads) * 100, 2)
                record = {
                    'agent': agent.full_name, 
                    'leads_count': (sb.tot_leads - sb.tot_deals),
                    'deals_count': sb.tot_deals,
                    'recycled_leads_count': 0,
                    'recycled_deals_count': 0,
                    'recycled_closing_percent': 0,
                    'total_leads': sb.tot_leads,
                    'total_closing_percent': closing_percent,
                    'retention': 0,
                    'total_debt': total_debt,
                }
                records.append(record)

            return records

        except Exception as err:
            app.logger.warning("Sales report service {}".format(str(err)))
            return []


from app.main.model.debt_payment import DebtPaymentSchedule, DebtEftStatus
class ACHReportSvc(ReportService):
    _heading = 'ACH Report'
    _report_name = 'ach_report.xlsx'
    
    @property
    def _fields(self):
        result = [ 'EPPS EFT Transaction ID', 'Payment Processor', 'Created Date', 'Client ID', 'Amount ($)', 
                  'Description', 'Effective Date', 'EPPS EFT Status', 'EPPS EFT Status Detail', 
                  'EPPS EFT Status Date', 'EPPS Trust Account Balance ($)', 'EPPS Account Holder ID', 
                  'Backend', 'Name Account Owner', 'Routing#', 'Account#', 'Account Type', 'Payment Transaction Id']
        return result
    
    def _records(self, data):
        try:
            records = []
            payment_records = DebtPaymentSchedule.query.filter(DebtPaymentSchedule.status != DebtEftStatus.FUTURE.name).all()
            for payment_record in payment_records:
                # bank account
                client = payment_record.contract.client
                bank_account = client.bank_account
                tr = payment_record.transaction

                record = {
                    'eft_trans_id': tr.trans_id if tr else "",
                    'pymt_processor': 'EPPS',
                    'created_date': self._dt2str(tr.created_date) if tr else "",
                    'client_id': "{} ({})".format(client.friendly_id, client.full_name),
                    'amount': payment_record.amount,
                    'description': payment_record.desc,
                    'effective_date': self._dt2str(payment_record.due_date),
                    'eft_status': tr.status if tr else "",
                    'eft_status_detail': tr.message if tr else "",
                    'eft_status_date': self._dt2str(tr.modified_date) if tr else "",
                    'trust_account_balance': 0,
                    'account_holder_id': client.epps_account_id,
                    'backend': 'EDMS',
                    'bank_name': bank_account.bank_name,
                    'routing_number': bank_account.routing_number,
                    'account_number': bank_account.account_number,
                    'account_type': 'Checking',
                    'pymt_trans_id': "",  
                }
                records.append(record)

            return records
        
        except Exception as err:
            app.logger.warning("ACH report service {}".format(str(err)))
            return []

class FutureDraftReportSvc(ReportService):
    _heading = 'Future Draft Transactions Report'
    _report_name = 'future_draft_report.xlsx'
    
    @property
    def _fields(self): 
        return ['Client ID', 'Amount($)', 'Description', 'Effective Date', 'Backend']

    def _records(self, data):
        try:
            records = []
            payment_records = DebtPaymentSchedule.query.filter(DebtPaymentSchedule.status == DebtEftStatus.FUTURE.name).all()
            for payment_record in payment_records:
                client = payment_record.contract.client
                friendly_id = client.friendly_id if client.friendly_id else ''
                record = {
                    'client_id': "{}({})".format(friendly_id, client.full_name),
                    'amount': payment_record.amount,
                    'description': payment_record.desc,
                    'effective_date': self._dt2str(payment_record.due_date),
                    'backend': 'EDMS',
                }
                records.append(record)

            return records

        except Exception as err:
            app.logger.warning("Future Draf report service {}".format(str(err)))
            return []

from app.main.model.collector import DebtCollector
class CollectorReportSvc(ReportService):
    _heading = 'Debt Collector Report'
    _report_name = 'debt_collector_report.xlsx'

    @property
    def _fields(self):
        return [ 'Collector Name', 'Company Name', 'Contact Phone #', 'Fax Number', 
                 'Number of Debt', 'Created Date', 'Status', 'Notes', 'Last Update' ]

    def _records(self, data):
        try:
            records = []
            query = DebtCollector.query

            collectors = query.all() 
            for collector in collectors:
                num_debts = collector.active_debts.count()
                record = {
                    'name': collector.name,
                    'company_name': '',
                    'phone': collector.phone,
                    'fax': collector.fax,
                    'num_debts': num_debts,
                    'created_date': self._dt2str(collector.inserted_on),
                    'status': 'Active',
                    'notes': '',
                    'last_update': self._dt2str(collector.updated_on),  
                }
                records.append(record)

            return records

        except Exception as err:
            app.logger.warning("Collector report service {}".format(str(err)))
            return []

from app.main.model.creditor import Creditor
class CreditorReportSvc(ReportService):
    _heading = 'Creditor Management Report'
    _report_name = 'creditors_report.xlsx'

    @property
    def _fields(self):
        return [ 'Creditor Name', 'Company Name', 'Contact Person', 'Contact Phone #', 
                 'Created Date', 'Status', 'Last Update' ]

    def _records(self, data):
        try:
            records = []
            query = Creditor.query

            creditors = query.all() 
            for creditor in creditors:
                record = {
                    'name': creditor.name,
                    'company_name': creditor.company_name,
                    'contact_person': creditor.contact_person,
                    'phone': creditor.phone,
                    'created_date': self._dt2str(creditor.inserted_on),
                    'status': 'Active' if creditor.is_active else 'Inactive',
                    'last_update': self._dt2str(creditor.updated_on),  
                }
                records.append(record)

            return records

        except Exception as err:
            app.logger.warning("Creditor report service {}".format(str(err)))
            return []

from app.main.model.usertask import UserTask
class TaskReportSvc(ReportService):
    _heading = 'Task Management Report'
    _report_name = 'task_report.xlsx'

    @property
    def _fields(self):
        return [ 'ID', 'Description', 'Status', 'Type', 'Client ID', 'Client Name', 
                 'Client Status', 'Account Manager', 'Team Manager', 'Due Date',
                 'Date Added' ]

    def _records(self, data):
        try:
            records = []
            filts = []
            if self._start:
                filts.append(func.date(UserTask.created_on)>=self._start)
            if self._end:
                filts.append(func.date(UserTask.created_on)<=self._end)

            query = UserTask.query 
            if len(filts) > 0:
                query = query.filter(and_(*filts))
            tasks = query.all()
            for task in tasks:
                client = task.client
                record = {
                    'id': task.id,
                    'desc': task.description,
                    'status': task.status,
                    'type': 'CS Task',
                    'client_id': client.friendly_id if client else '',
                    'client_name': client.full_name if client else '',
                    'client_status': client.disposition.value if client else '',
                    'account_manager': task.owner.full_name,
                    'team_manager': client.team_manager.full_name if client.team_manager else '',
                    'due_date': self._dt2str(task.due_date),
                    'date_added': self._dt2str(task.created_on), 
                }
                records.append(record)

            return records

        except Exception as err:
            app.logger.warning("Task report service {}".format(str(err)))
            return []

from app.main.model.credit_report_account import CreditReportData
class DaysDelinquentReportSvc(ReportService):
    _heading = 'Days Delinquent Report'
    _report_name = 'days_delinquent_report.xlsx'
   
    @property
    def _fields(self):
        return [ 'Credit Report Status', 'Days Delinquent', 'Last Update']

    def _records(self, data):
        try:
            records = []
            filts = []
            query = CreditReportData.query 
            if len(filts) > 0:
               query = query.filter(and_(*filts)) 
            debts = query.order_by(desc(CreditReportData.last_update)).all()
            for debt in debts:
                record = {
                    'status': '',
                    'days_delinquent': debt.days_delinquent,
                    'last_update': self._dt2str(debt.last_update),
                } 
                records.append(record)

            return records

        except Exception as err:
            app.logger.warning("Days Delinquent report service {}".format(str(err)))
            return []

from app.main.model.debt_payment import EftReturnFee
class EftReturnFeeReportSvc(ReportService):
    _heading = 'EFT Return Fee Report'
    _report_name = 'eft_return_fee_report.xlsx'

    @property
    def _fields(self):
        return [ 'Code', 'Fee Amount($)', 'Last Update' ]

    def _records(self, data):
        try:
            records = []
            filts = []
            query = EftReturnFee.query
            if len(filts) > 0:
               query = query.filter(and_(*filts))
            items = query.order_by(desc(EftReturnFee.modified_date)).all()
            for item in items:
                last_update_str = "{}, by {}".format(self._dt2str(item.modified_date),
                                                     item.agent.full_name)
                record = {
                    'code': item.code,
                    'amount': item.amount,
                    'last_update': last_update_str,
                }
                records.append(record)

            return records

        except Exception as err:
            app.logger.warning("Eft Return Fee report service {}".format(str(err)))
            return []

from app.main.model.template import MailBox, TemplateMedium
class PostMailReportSvc(ReportService):
    _heading = 'Post Mail Print Report'
    _report_name = 'post_mail_print_report.xlsx'

    @property
    def _fields(self):
        return [ 'Client Id', 'Client Status', 'Recipient Name', 'Recipient Address', 
                 'Recipient Phone', 'File Name', 'Sent Date', 'Date Added' ]

    def _records(self, data):
        try:
            records = []
            filts = []
            query = MailBox.query
            filts.append(MailBox.channel==TemplateMedium.POST.name)
            if len(filts) > 0:
               query = query.filter(and_(*filts))
            mails = query.order_by(desc(MailBox.timestamp)).all()
            for mail in mails:
                client = mail.client
                debt_collector = mail.debt_collector
                template = mail.template
                record = {
                    'client_id': client.id,
                    'client_status': client.disposition.value if client.disposition else '',
                    'recipient_name': debt_collector.name if debt_collector else '',
                    'recipient_addr': debt_collector.address if debt_collector else '',
                    'recipient_phone': debt_collector.phone if debt_collector else '',
                    'file_name': template.fname,
                    'sent_date': '',
                    'added_date': self._dt2str(mail.timestamp), 
                }

            return records
            
        except Exception as err:
            app.logger.warning("Post Mail print report service {}".format(str(err)))
            return []

from app.main.model.team import Team
class TeamReportSvc(ReportService):
    _heading = 'Team Report'
    _report_name = 'team_report.xlsx'

    @property
    def _fields(self):
        return [ 'Team Name', 'Account Manager', 'Team Manager', 'Status', 
                 'Description', 'Last Update' ]   

    def _records(self, data):
        try:
            records = []
            filts = []
            query = Team.query
            if len(filts) > 0:
                query = query.filter(and_(*filts))
            teams = query.order_by(desc(Team.modified_date)).all()
            for team in teams:
                manager = team.manager.full_name if team.manager else ''
                creator = team.creator.full_name if team.creator else ''
                tms = team.team_members.all()
                agent_list = [tm.member.full_name for tm in tms]
                record = {
                    'name': team.name,
                    'accout_manager': agent_list,
                    'team_manager': manager, 
                    'status': 'Active' if team.is_active else 'Inactive',
                    'description': team.description,
                    'last_update': creator,
                }
                records.append(record)
            
            return records

        except Exception as err:
            app.logger.warning("Post Mail print report service {}".format(str(err)))
            return []

class StaffReportSvc(ReportService):
    _heading = 'Staff Report'
    _report_name = 'staff_report.xlsx'

    @property
    def _fields(self):
        return [ 'Username', 'Full Name', 'Routed-Call Phone #', 'Ext', 'Inboud DID', 
                 'Phone', 'Sales Manager', 'Department', 'Status', 'Date of starting work', 
                 'Park', 'Comments', 'Last Update' ]

    def _records(self, data):
        try:
            records = []
            filts = []
            query = User.query
            if len(filts) > 0:
                query = query.filter(and_(*filts))
            users = query.order_by(desc(User.registered_on)).all()
            for user in users:
                team = ''
                if user.team_member:
                    team = user.team_member.team.name  
                record = {
                    'username': user.username,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'routed_call_phone': user.voip_route_number,
                    'ext': '',
                    'inbound_did': '',
                    'phone_number': user.personal_phone,
                    'sales_manager': '',
                    'department': team,
                    'status': 'Active',
                    'start_date': self._dt2str(user.start_date), # add field in table 
                    'park': '',
                    'comments': '',
                    'last_update': "{} by Admin".format(self._dt2str(user.modified_date)), 
                }
                records.append(record)

            return records

        except Exception as err:
            app.logger.warning("Staff Management report service {}".format(str(err)))
            return []

from app.main.model.debt import DebtDispute
class DebtMgmtReportSvc(ReportService):
    _heading = 'Debt Management Report'
    _report_name = 'debt_mgmt_report.xlsx'

    @property
    def _fields(self):
        return [ 'Client ID', 'Client Full Name', 'P1 Received Date', 'P2 Received Date', 
                 'NOIR/NOIR2/NOIR FDCPA Received Date', 'SP Received Date', 'Fully Disputed Date', 
                 'Notes', 'Client Status', 'Account Manager', 'Debt ID', 'Debt Name', 
                 'ECOA', 'Creditor', 'Type', 'Account Number', 'Push' ] 


    def _records(self, data):
        try:
            records = []
            filts = []
            query = DebtDispute.query
            if len(filts) > 0:
                query = query.filter(and_(*filts))
            debt_disputes = query.order_by(desc(DebtDispute.modified_date)).all()
            for debt_dispute in debt_disputes:
                debt = debt_dispute.debt
                client = debt_dispute.client
                record = {
                    'client_id': debt_dispute.client.friendly_id,
                    'client_name': debt_dispute.client.full_name,
                    'p1_date': self._dt2str(debt_dispute.p1_date),
                    'p2_date': self._dt2str(debt_dispute.p2_date),
                    'noir_date': self._dt2str(debt_dispute.noir_date),
                    'sp_date': self._dt2str(debt_dispute.sp_date),
                    'fully_disputed_date': self._dt2str(debt_dispute.fully_disputed_date),
                    'notes': '',
                    'client_status': client.disposition.value,
                    'account_manager': client.account_manager.full_name if client.account_manager else '',
                    'debt_id': debt.public_id,
                    'debt_name': debt.debt_name,
                    'ecoa': debt.ecoa,
                    'creditor': debt.creditor,
                    'debt_type': debt.account_type,
                    'debt_account_no': debt.account_number,
                    'push': 'Yes' if debt.push else 'No',  
                }
                records.append(record)
           
            return records

        except Exception as err:
            app.logger.warning("Staff Management report service {}".format(str(err)))
            return []

# Notification Reports
class NotificationReportSvc(ReportService):
    _heading = 'Notification Management Report'
    _report_name = 'notification_mgmt_report.xlsx'

    @property
    def _fields(self):
        return [ 'Notification Date', 'Notify via', 'Notify to Email/Phone #', 'Content',
                 'Type', 'Status', 'Client ID', 'Client Full Name' ]

    def _records(self, data):
        try:
            records = []
            filts = [] 

            return records

        except Exception as err:
            app.logger.warning("Staff Management report service {}".format(str(err)))
            return []

# Service Status Report
class StatusReportSvc(ReportService):
    _heading = 'Status Report'
    _report_name = 'status_report.xlsx'

    @property
    def _fields(self):
        return [ 'Client ID', 'Client Name', 'Enrolled Date', 'Account Status' 'Debt Name', 
                 'Push', 'Days Delinquent', 'Debt Status', 'Service Status', 'Next CS Schedule Apt', 
                 'Times of Reschedule', 'Account Manager', 'Backend', 'Attorney', 'Sales Date', 
                 'Next Payment Date' ]

    def _records(self, data):
        try:
            records = []
            filts = []
            query = CreditReportData.query
            if len(filts) > 0:
                query = query.filter(and_(*filts))
            debts = query.order_by(desc(CreditReportData.enrolled_date)).all()
            for debt in debts:
                credit_account = debt.credit_report_account
                if not credit_account:
                    continue
                client = credit_account.client
                diff = (datetime.now() - debt.enrolled_date)
                days_delinquent = 0
                if debt.days_delinquent:
                    days_delinquent = int(debt.days_delinquent) + diff.days
                
                record = {
                    'client_id': client.friendly_id,
                    'client_name': client.full_name,
                    'enrolled_date': self._dt2str(debt.enrolled_date),
                    'account_status': client.disposition.value if client.disposition else '',
                    'debt_name': debt.debt_name,
                    'push': 'Yes' if debt.push else '',
                    'days_delinquent': days_delinquent,
                    'debt_status': '',
                    'service_status': '',
                    'next_service_schedule_apt': '',
                    'times_reschedule': 0,
                    'account_manager': client.account_manager.full_name if client.account_manager else '',
                    'backend': 'EDMS',
                    'attorney': '',
                    'sales_date': '',
                    'next_payment_date': '',
                }
                records.append(record)

            return records

        except Exception as err:
            app.logger.warning("Status report service {}".format(str(err)))
            return []

# Service Status Report
class AccountAdminReportSvc(ReportService):
    _heading = 'Account Admin Report'
    _report_name = 'account_admin_report.xlsx'

    @property
    def _fields(self):
        return [ 'Account Tranaction ID', 'Transaction Date', 'Post Date', 'Client ID' 
                 'Amount', 'Balance', 'For/From', 'Description',
               ]

    
    def _records(self, data):
        try:
            records = []
            filts = []
            return records
    
        except Exception as err:
            app.logger.warning("Account Admin report service {}".format(str(err)))
            return []

