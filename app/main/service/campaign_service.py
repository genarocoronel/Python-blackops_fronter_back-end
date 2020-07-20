from datetime import datetime
import uuid

from app.main.model.campaign import Campaign, MarketingModel, MailType, PinnaclePhoneNumber
from app.main.model.client import Client, ClientType, ClientCampaign
from sqlalchemy import and_
from .apiservice import ApiService, has_permissions

class CampaignService(ApiService):
    _model = Campaign
    _key_field   = 'public_id'
    
    def _parse(self, data, insert=True):
        ## validate if necessary
        mkt_type = data.get('marketing_type')
        mm = MarketingModel[mkt_type]
        mail_type = data.get('mail_type')
        mt = MailType[mail_type]
        name = data.get('name')
        num_pieces = data.get('num_mail_pieces')

        min_debt = data.get('min_debt')
        max_debt = data.get('max_debt')
        debt_range = {
            'min': min_debt if min_debt else 0,
            'max': max_debt if max_debt else 0,
        }

           
        pin_phone_no = data.get('pinnacle_phone')
        ppn = PinnaclePhoneNumber.query.filter_by(number=pin_phone_no).first()
        if not ppn:
            # should we create new entry ??
            raise ValueError("Pinnacle Phone Number not present")

        fields = {
            'name': name,
            'description': data.get('description'),
            'phone': data.get('phone'),
            'offer_expire_date': data.get('offer_expire_date'),
            'mailing_date': data.get('mailing_date'),
            'pinnacle_phone_num_id': ppn.id, 
            'marketing_model': mm.name,
            'mail_type': mt.name, 
            'num_mail_pieces': num_pieces,
            'cost_per_piece': data.get('cost_per_piece'),
            'est_debt_range': debt_range,
        }
        if insert is True:
            mail_dtime = datetime.strptime(data.get('mailing_date'), '%Y-%m-%d')
            month_word = mail_dtime.strftime("%B")
            job_num = f'{mkt_type}_{month_word}{mail_dtime.day}{mail_dtime.year}_{mail_type}_{num_pieces}_{min_debt}-{max_debt}'

            ## insert fields
            fields['public_id'] = str(uuid.uuid4())
            fields['job_number'] = job_num
            fields['inserted_on'] = datetime.utcnow(),

        return fields

    def _validate(self, fields, method='save'): 
        if method == 'save':
            obj = self._model.query.filter_by(name=fields.get('name')).first()     
            if obj:
                raise ValueError("Campaign name already exists in the database.")
    

    def _queryset(self):
        return self._model.query.all()

class PinnaclePhoneNumService(ApiService):
    _model = PinnaclePhoneNumber
    
    def _queryset(self):
        return self._model.query.all()

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from app.main.core.io import stream_file
from app.main.config import upload_location as report_file_path

class CampaignReportService(ApiService):
    _heading = 'DIRECT MAIL ZIP PERFORMANCE SUMMARY'
    _dealer  = 'ELITE DMS'
    _report_name = 'campaign_report.xlsx'
    
    @has_permissions
    def generate_report(self):
        wb = Workbook()
        ws = wb.active

        # fill the data
        ws['A1'] = self._heading
        ws['A3'] = 'Dealer Name:'
        ws['C3'] = self._dealer
        ws['A4'] = 'Location:'
        ws['A5'] = 'Report Date:'
        ws['C5'] = datetime.utcnow().strftime('%m/%d/%Y')
        # styling
        ws['A1'].font = Font(bold=True)

        ws.append([])
        ws.append(['NAME OF CAMPAIGN', 'MAIL PIECES SENT', 'LEADS GENERATED', 'LEAD %', 'DEALS', 'DEAL %', 'TOTAL COST', 'COST/LEAD', 'COST/DEAL'])
        col_head_style = PatternFill(start_color='0591f5', end_color='0591f5', fill_type = "solid")
        for rows in ws.iter_rows(min_row=7, max_row=7, min_col=1):
            for cell in rows:
                cell.fill = col_head_style
         
        report = []
        row_index = 7
        for campaign in Campaign.query.all():
            row = []
            row_index = row_index + 1
            if not campaign.num_mail_pieces:
                campaign.num_mail_pieces = 0
            if not campaign.cost_per_piece:
                campaign.cost_per_piece = 0

            # leads generated
            leads_percent = 0 
            num_leads = ClientCampaign.query.outerjoin(Client)\
                                            .outerjoin(Campaign)\
                                            .filter(and_(Campaign.id==campaign.id, Client.type==ClientType.lead)).count()

            if num_leads > 0:
                leads_percent = round((num_leads / campaign.num_mail_pieces) * 100, 2)
            # deals
            deals_percent = 0
            num_clients = ClientCampaign.query.outerjoin(Client)\
                                              .outerjoin(Campaign)\
                                              .filter(and_(Campaign.id==campaign.id, Client.type==ClientType.client)).count()
            if num_clients > 0:
                deals_percent = round((num_clients / campaign.num_mail_pieces) * 100, 2) 
            total_cost = campaign.num_mail_pieces * campaign.cost_per_piece 

            cost_per_lead = round(total_cost / num_leads, 2) if num_leads > 0 else 0
            cost_per_deal = round(total_cost / num_clients, 2) if num_clients > 0 else 0

            row.append(campaign.name)
            row.append(campaign.num_mail_pieces)
            row.append(num_leads)
            row.append(leads_percent)
            row.append(num_clients)
            row.append(deals_percent)
            row.append(total_cost)
            row.append(cost_per_lead)
            row.append(cost_per_deal)
            ws.append(row) 
        
        t = row_index + 1 
        ws["B{}".format(t)] = '=SUM(B8:B{})'.format(row_index)
        ws["C{}".format(t)] = '=SUM(C8:C{})'.format(row_index)
        ws["D{}".format(t)] = '=(C{}/B{})*100'.format(t, t)
        ws["E{}".format(t)] = '=SUM(E8:E{})'.format(row_index)
        ws["G{}".format(t)] = '=SUM(G8:G{})'.format(row_index)
        # save the file    
        report_file = "{}/{}".format(report_file_path, self._report_name)  
        wb.save(filename=report_file)

        return stream_file(report_file_path, self._report_name, as_attachment=False)

        

